import os
import csv
import subprocess
import openpyxl
from pypdf import PdfReader

class ExtractorDeArchivos:
    """
    Servicio de extracción agnóstico y multiplataforma.
    Toma un archivo, identifica su extensión y extrae el texto puro,
    estructuras (tablas convertidas a texto) o metadatos/audio de multimedia.
    """
    def __init__(self):
        pass

    def procesar(self, ruta_archivo: str):
        """
        Punto de entrada principal. Identifica el archivo y extrae su contenido.
        """
        ruta_archivo = os.path.normpath(os.path.abspath(ruta_archivo))
        if not os.path.exists(ruta_archivo):
            raise FileNotFoundError(f"El archivo no existe: {ruta_archivo}")
            
        ext = os.path.splitext(ruta_archivo)[1].lower()
        
        # 1. ARCHIVOS DE TEXTO PLANO Y DOCUMENTOS
        if ext in ['.txt', '.md', '.json', '.xml']:
            return self._extraer_texto_plano(ruta_archivo)
            
        elif ext == '.pdf':
            return self._extraer_pdf(ruta_archivo)
            
        # 2. ARCHIVOS DE DATOS / TABLAS
        elif ext in ['.xlsx', '.xls']:
            return self._extraer_excel(ruta_archivo)
            
        elif ext == '.csv':
            return self._extraer_csv(ruta_archivo)
            
        # 3. ARCHIVOS MULTIMEDIA (VIDEO/AUDIO)
        elif ext in ['.mp4', '.avi', '.mkv', '.mov', '.mp3', '.wav', '.m4a']:
            return self._procesar_multimedia(ruta_archivo)
            
        else:
            return f"[Error: Formato {ext} no soportado actualmente]"

    # ==========================================
    # LÓGICAS INTERNAS DE EXTRACCIÓN
    # ==========================================

    def _extraer_texto_plano(self, ruta):
        with open(ruta, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()

    def _extraer_pdf(self, ruta):
        lector = PdfReader(ruta)
        texto_completo = []
        for i, pagina in enumerate(lector.pages):
            texto_pag = pagina.extract_text()
            if texto_pag:
                texto_completo.append(f"[Inicio Página {i+1}]\n{texto_pag}\n[Fin Página {i+1}]")
        return "\n".join(texto_completo)

    def _extraer_excel(self, ruta):
        """Transforma las tablas en texto semántico para que el LLM entienda las filas"""
        wb = openpyxl.load_workbook(ruta, data_only=True)
        resultado = []
        
        for nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
            resultado.append(f"\n## Hoja de cálculo: {nombre_hoja} ##")
            
            for fila in hoja.iter_rows(values_only=True):
                # Filtrar filas completamente vacías
                if any(celda is not None for celda in fila):
                    # Convertir cada celda a string y unirlas con un separador visual limpio (|)
                    fila_limpia = " | ".join([str(celda).strip() if celda is not None else "" for celda in fila])
                    resultado.append(fila_limpia)
        return "\n".join(resultado)

    def _extraer_csv(self, ruta):
        resultado = []
        with open(ruta, mode='r', encoding='utf-8', errors='ignore') as f:
            # Detección de dialecto básica para CSVs de Windows/Linux (delimitador , o ;)
            muestra = f.read(1024)
            f.seek(0)
            try:
                dialecto = csv.Sniffer().sniff(muestra)
                lector = csv.reader(f, dialecto)
            except csv.Error:
                lector = csv.reader(f)
                
            for fila in lector:
                if any(celda.strip() for celda in fila):
                    resultado.append(" | ".join(fila))
        return "\n".join(resultado)

    def _procesar_multimedia(self, ruta):
        """
        Para multimedia se devuelve un diccionario con rutas a los recursos
        separados (audio wav para el STT, etc).
        """
        datos_multimedia = {
            "tipo": "multimedia",
            "ruta_original": ruta,
            "ruta_audio_extraido": None,
            "fotogramas_clave": []
        }
        
        # Extracción de audio multiplataforma usando subprocess para mayor seguridad
        ruta_audio = os.path.splitext(ruta)[0] + "_audio_extract.wav"
        
        try:
            # Se usa subprocess en lugar de os.system para evitar problemas con espacios en rutas en Windows/Linux
            # Se asume que ffmpeg está instalado y en el PATH del sistema
            cmd = [
                "ffmpeg", "-y", "-i", ruta,
                "-ab", "160k", "-ac", "2", "-ar", "44100", "-vn",
                ruta_audio
            ]
            
            # Ejecutar silenciosamente
            subprocess.run(
                cmd, 
                stdout=subprocess.DEVNULL, 
                stderr=subprocess.DEVNULL, 
                check=True
            )
            datos_multimedia["ruta_audio_extraido"] = ruta_audio
        except Exception as e:
            # Fallback seguro en caso de error (ej. ffmpeg no instalado)
            datos_multimedia["error"] = f"No se pudo extraer audio. Requiere ffmpeg en PATH. Detalles: {str(e)}"
            
        return datos_multimedia

# Bloque de prueba standalone y CLI
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        extractor = ExtractorDeArchivos()
        arch = sys.argv[1]
        
        try:
            datos = extractor.procesar(arch)
            
            # Si nos pasan un segundo argumento, lo tomamos como archivo de salida
            if len(sys.argv) > 2:
                out_path = sys.argv[2]
                if isinstance(datos, str):
                    with open(out_path, 'w', encoding='utf-8') as f:
                        f.write(datos)
                else:
                    import json
                    with open(out_path, 'w', encoding='utf-8') as f:
                        json.dump(datos, f, ensure_ascii=False, indent=2)
            else:
                print(f"\n--- Procesando: {arch} ---")
                if isinstance(datos, str):
                    print(f"Texto extraído ({len(datos)} caracteres):\n{datos[:200]}...")
                else:
                    print(f"Datos estructurados:\n{datos}")
        except Exception as e:
            print(f"Error procesando {arch}: {e}")

